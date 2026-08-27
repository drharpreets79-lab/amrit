#!/usr/bin/env python3
"""Generate the bundled EUCAST breakpoint extract.

Why this exists
---------------
CLSI's M100 is a paid standard, so AMRIT can only ever link to it. EUCAST publishes its
clinical breakpoint tables free of charge and permits redistribution with attribution, which
means a deployment that reads EUCAST can ship with real breakpoints already installed rather
than an empty Breakpoint Centre and a note telling a laboratory to type them in.

The application's "Update EUCAST breakpoints" button fetches the current table directly. This
script produces the offline fallback for the very common case of a laboratory with no
outbound network: run it on a connected machine, commit the result, and the installer carries
it.

Nothing here invents a threshold. Every value comes from EUCAST's own workbook; if the
download fails or the layout is not recognised, the script exits non-zero and writes nothing,
because a breakpoint file with made-up numbers in it is worse than no file at all.

Usage
-----
    python3 tools/fetch_eucast_breakpoints.py                  # newest known edition
    python3 tools/fetch_eucast_breakpoints.py --version 16.0
    python3 tools/fetch_eucast_breakpoints.py --from-file v_16.0_Breakpoint_Tables.xlsx
    python3 tools/fetch_eucast_breakpoints.py --check          # verify the checked-in file

Output: app/resources/breakpoints/eucast-breakpoints.json
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
import urllib.request
from datetime import datetime, timezone
from pathlib import Path

BASE_URL = "https://www.eucast.org/fileadmin/src/media/PDFs/EUCAST_files/Breakpoint_tables/"
# Newest first. EUCAST keeps no "latest" alias, so a new edition is added here.
KNOWN_VERSIONS = ("16.0", "15.0", "14.0", "13.1", "13.0")
OUTPUT = Path(__file__).resolve().parent.parent / "app" / "resources" / "breakpoints" / "eucast-breakpoints.json"

ATTRIBUTION = (
    "European Committee on Antimicrobial Susceptibility Testing. Breakpoint tables for "
    "interpretation of MICs and zone diameters. Published free of charge by EUCAST; "
    "redistributed with attribution."
)

S_MIC = re.compile(r"^s\s*(?:≤|<=)", re.I)
R_MIC = re.compile(r"^r\s*(?:>)", re.I)
S_ZONE = re.compile(r"^s\s*(?:≥|>=)", re.I)
R_ZONE = re.compile(r"^r\s*(?:<)", re.I)
# A breakpoint is a number. EUCAST writes guidance ("Note", "IE", "-") where a value would
# go, and repeats its banded heading once per drug class, so the heading text ("MIC
# breakpoints (mg/L)") lands in the value column of a row whose agent is a class name.
THRESHOLD = re.compile(r"^\(?\s*(?:≤|≥|<|>|<=|>=)?\s*=?\s*(\d+(?:[.,]\d+)?)\s*\)?$")
BRACKETED = re.compile(r"^\(.*\)$")
# EUCAST sub-divides the anaerobe sheet by organism, naming each band with a bare label.
PROSE = re.compile(r"\b(see|breakpoints?|valid|based|note|dosage|table|used?|are|is|for|with|the)\b", re.I)
# Sheets that are not organism tables. "Guidance" matters most: its worked examples are
# attributed to invented agents ("Antimicrobial agent A") and read exactly like breakpoints.
NON_TABLE_SHEETS = {"content", "contents", "changes", "notes", "guidance", "dosages", "technical uncertainty"}
PLACEHOLDER_AGENT = re.compile(r"^antimicrobial agent\b", re.I)


def table_url(version: str) -> str:
    return f"{BASE_URL}v_{version}_Breakpoint_Tables.xlsx"


def download(version: str) -> bytes:
    url = table_url(version)
    request = urllib.request.Request(url, headers={"User-Agent": "AMRIT-breakpoint-fetch/1.0"})
    with urllib.request.urlopen(request, timeout=120) as response:  # noqa: S310 - fixed https host
        if response.status != 200:
            raise RuntimeError(f"HTTP {response.status} for {url}")
        return response.read()


def cell(value: object) -> str:
    """A cell's text with EUCAST's footnote markers removed.

    EUCAST attaches footnotes as superscript digits inside the cell, so an
    ampicillin-sulbactam breakpoint of 2 mg/L carrying footnote 1 flattens to "21" — a
    ten-fold error in a number that decides whether an isolate is reported susceptible.
    The workbook keeps the run formatting, so superscript runs are dropped and only the
    value's own runs are read. No heuristic could do this over the flattened text: "21" is
    a perfectly ordinary zone diameter.
    """
    if value is None:
        return ""
    from openpyxl.cell.rich_text import CellRichText, TextBlock

    if isinstance(value, CellRichText):
        kept = []
        for part in value:
            if isinstance(part, TextBlock):
                font = getattr(part, "font", None)
                if font is not None and getattr(font, "vertAlign", None) == "superscript":
                    continue
                kept.append(str(part.text))
            else:
                kept.append(str(part))
        return "".join(kept).strip()
    return str(value).strip()


def threshold(value: str) -> tuple[str, bool] | None:
    """The number in a value cell, and whether EUCAST bracketed it."""
    text = (value or "").strip()
    if not text:
        return None
    matched = THRESHOLD.match(text)
    if not matched:
        return None
    return matched.group(1).replace(",", "."), bool(BRACKETED.match(text))


def looks_like_organism(label: str) -> bool:
    text = (label or "").strip()
    if not text or len(text) > 60 or PLACEHOLDER_AGENT.match(text):
        return False
    if PROSE.search(text):
        return False
    return len(text.split()) <= 5


def parse(workbook_bytes: bytes, version: str) -> list[dict[str, str]]:
    """Read every organism sheet. Mirrors parseEucastTables in app/src/main/services.ts."""
    try:
        from openpyxl import load_workbook
    except ImportError:  # pragma: no cover - a maintainer-machine dependency
        raise SystemExit("openpyxl is required: pip install openpyxl") from None

    import io

    workbook = load_workbook(io.BytesIO(workbook_bytes), read_only=True, data_only=True, rich_text=True)
    rows: list[dict[str, str]] = []
    for sheet_name in workbook.sheetnames:
        if sheet_name.strip().lower() in NON_TABLE_SHEETS:
            continue
        sheet = workbook[sheet_name]
        matrix = [[cell(value) for value in row] for row in sheet.iter_rows(max_row=400, values_only=True)]
        # Every band on a sheet repeats the header. Reading only the first one treated the
        # remaining headings as agents and, on the anaerobe sheet, filed four organisms'
        # breakpoints under one name.
        header_indexes = [
            index for index, line in enumerate(matrix)
            if any(S_MIC.match(c) for c in line) and any(R_MIC.match(c) for c in line)
        ]
        for band, header_index in enumerate(header_indexes):
            header = matrix[header_index]
            columns = {
                "s_mic": next((i for i, c in enumerate(header) if S_MIC.match(c)), -1),
                "r_mic": next((i for i, c in enumerate(header) if R_MIC.match(c)), -1),
                "s_zone": next((i for i, c in enumerate(header) if S_ZONE.match(c)), -1),
                "r_zone": next((i for i, c in enumerate(header) if R_ZONE.match(c)), -1),
            }
            above = matrix[header_index - 1] if header_index > 0 else []
            band_label = (header[0] or (above[0] if above else "")).strip()
            drug_class = "" if PLACEHOLDER_AGENT.match(band_label) else band_label
            organism_name = sheet_name.strip()
            if not drug_class:
                floor = header_indexes[band - 1] + 1 if band > 0 else 0
                for index in range(header_index - 1, floor - 1, -1):
                    candidate = matrix[index]
                    if not candidate or not candidate[0].strip() or any(c.strip() for c in candidate[1:]):
                        continue
                    if looks_like_organism(candidate[0]):
                        organism_name = candidate[0].strip()
                        break
            end = max(header_index + 1, header_indexes[band + 1] - 1) if band + 1 < len(header_indexes) else len(matrix)

            for line in matrix[header_index + 1:end]:
                agent = line[0].strip() if line else ""
                if not agent or len(agent) > 120 or PLACEHOLDER_AGENT.match(agent):
                    continue
                get = lambda key: line[columns[key]] if 0 <= columns[key] < len(line) else ""  # noqa: E731
                mic_s, mic_r = threshold(get("s_mic")), threshold(get("r_mic"))
                zone_s, zone_r = threshold(get("s_zone")), threshold(get("r_zone"))
                tail = max(columns["r_mic"], columns["r_zone"]) + 1
                bracketed = any(item[1] for item in (mic_s, mic_r, zone_s, zone_r) if item)
                comments = "; ".join(part for part in (
                    f"EUCAST section: {drug_class}" if drug_class else "",
                    "EUCAST published this threshold in brackets; check the edition notes before reporting it."
                    if bracketed else "",
                    " ".join(c for c in line[tail:] if c),
                ) if part)[:500]
                if mic_s or mic_r:
                    rows.append({
                        "guideline": "EUCAST", "edition": version, "test_method": "MIC",
                        "antibiotic_code": "", "antibiotic_name": agent,
                        "organism_code": "", "organism_name": organism_name,
                        "susceptible": mic_s[0] if mic_s else "",
                        "intermediate": "",
                        "resistant": mic_r[0] if mic_r else "",
                        "units": "mg/L", "comments": comments, "source_sheet": sheet_name,
                        "fda_susceptible": "", "fda_intermediate": "", "fda_resistant": "", "clsi_fda_match": "",
                    })
                if zone_s or zone_r:
                    rows.append({
                        "guideline": "EUCAST", "edition": version, "test_method": "Disk diffusion",
                        "antibiotic_code": "", "antibiotic_name": agent,
                        "organism_code": "", "organism_name": organism_name,
                        "susceptible": zone_s[0] if zone_s else "",
                        "intermediate": "",
                        "resistant": zone_r[0] if zone_r else "",
                        "units": "mm", "comments": comments, "source_sheet": sheet_name,
                        "fda_susceptible": "", "fda_intermediate": "", "fda_resistant": "", "clsi_fda_match": "",
                    })
    return rows


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--version", help="EUCAST edition, e.g. 16.0. Defaults to the newest that downloads.")
    parser.add_argument("--from-file", type=Path, help="Use a workbook already on disk instead of downloading.")
    parser.add_argument("--check", action="store_true", help="Verify the checked-in extract without writing.")
    args = parser.parse_args()

    if args.check:
        if not OUTPUT.exists():
            print(f"No bundled extract at {OUTPUT}. The in-app update button still works online.")
            return 0
        payload = json.loads(OUTPUT.read_text())
        digest = hashlib.sha256(json.dumps(payload["rows"], sort_keys=True).encode()).hexdigest()
        ok = digest == payload.get("rows_sha256")
        print(f"{OUTPUT}: v{payload.get('version')} · {len(payload['rows'])} rows · checksum {'ok' if ok else 'MISMATCH'}")
        return 0 if ok else 1

    if args.from_file:
        version = args.version or (re.search(r"v_([\d.]+)_", args.from_file.name) or ["", "unknown"])[1]
        blob = args.from_file.read_bytes()
    else:
        candidates = [args.version] if args.version else list(KNOWN_VERSIONS)
        blob, version = None, ""
        for candidate in candidates:
            try:
                blob = download(candidate)
                version = candidate
                print(f"Downloaded EUCAST v{candidate}")
                break
            except Exception as error:  # noqa: BLE001 - report and try the previous edition
                print(f"  v{candidate}: {error}", file=sys.stderr)
        if blob is None:
            print("No EUCAST edition could be downloaded; nothing written.", file=sys.stderr)
            return 1

    rows = parse(blob, version)
    if not rows:
        print("No breakpoint rows recognised. EUCAST changed the layout; update the parser "
              "here and in app/src/main/services.ts before shipping anything.", file=sys.stderr)
        return 1

    payload = {
        "version": version,
        "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "source_url": table_url(version),
        "source_sha256": hashlib.sha256(blob).hexdigest(),
        "attribution": ATTRIBUTION,
        "rows_sha256": hashlib.sha256(json.dumps(rows, sort_keys=True).encode()).hexdigest(),
        "rows": rows,
    }
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(payload, indent=1, ensure_ascii=False) + "\n")
    print(f"Wrote {len(rows)} rows to {OUTPUT}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
